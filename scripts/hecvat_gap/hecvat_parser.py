"""Extract vendor/product/date fields from HECVAT xlsx attachments.

HECVAT templates vary by version and by whatever a vendor customized, so
this scans every cell for known field labels instead of assuming fixed
cell references, and takes the first non-empty neighbor (same row to the
right, else the cell directly below) as that field's value.
"""
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl

LABEL_PATTERNS = {
    "vendor": re.compile(r"vendor\s*(legal)?\s*name|company\s*name", re.I),
    "product": re.compile(r"product\s*(or|/)?\s*service\s*name|application\s*name|product\s*name", re.I),
    "date": re.compile(r"date\s*(completed|submitted|of\s*completion)?\s*$|completion\s*date", re.I),
}

TICKET_NUMBER_RE = re.compile(r"(RITM|INC|REQ|SCTASK)\d+", re.I)

DATE_FORMATS = ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y", "%d-%b-%Y")

# Words that show up in downloaded HECVAT filenames but aren't the vendor name,
# stripped out when falling back to guessing the vendor from the filename.
FILENAME_STOPWORDS = {
    "hecvat", "lite", "full", "assessment", "questionnaire", "vendor", "final", "signed", "v1", "v2", "v3", "copy",
}


@dataclass
class HecvatRecord:
    source_file: str
    ticket_number: str | None
    vendor: str | None
    product: str | None
    assessment_date: datetime | None
    vendor_source: str | None = None  # "cell" or "filename" -- lets callers weight confidence
    issues: list = field(default_factory=list)


def _find_value(ws, label_pattern):
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not label_pattern.search(cell.value.strip()):
                continue
            for other in row[cell.column :]:
                if other.value not in (None, ""):
                    return other.value
            below = ws.cell(row=cell.row + 1, column=cell.column)
            if below.value not in (None, ""):
                return below.value
    return None


def _clean(value):
    return str(value).strip() if value not in (None, "") else None


def _coerce_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if hasattr(value, "year") and hasattr(value, "month"):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def _vendor_from_filename(path: Path) -> str | None:
    stem = TICKET_NUMBER_RE.sub("", path.stem)
    words = [w for w in re.split(r"[_\-. ]+", stem) if w and w.lower() not in FILENAME_STOPWORDS]
    guess = " ".join(words).strip()
    return guess or None


def parse_file(path: Path) -> HecvatRecord:
    ticket_match = TICKET_NUMBER_RE.search(path.stem)
    ticket_number = ticket_match.group(0).upper() if ticket_match else None

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return HecvatRecord(str(path), ticket_number, None, None, None, issues=[f"could not open: {exc}"])

    vendor = product = date_val = None
    for ws in wb.worksheets:
        vendor = vendor or _find_value(ws, LABEL_PATTERNS["vendor"])
        product = product or _find_value(ws, LABEL_PATTERNS["product"])
        date_val = date_val or _find_value(ws, LABEL_PATTERNS["date"])
    wb.close()

    vendor = _clean(vendor)
    issues = []
    vendor_source = "cell" if vendor else None

    if not vendor:
        # No "Vendor Name"-style field found in the file itself -- fall back to
        # guessing from the downloaded filename (e.g. "RITM0012345_Zoom.xlsx").
        # This is a real gap some HECVAT exports hit, not just a hypothetical:
        # some org's exported questionnaires omit a labeled vendor field, or
        # the tickets/attachments simply don't carry it at all.
        filename_guess = _vendor_from_filename(path)
        if filename_guess:
            vendor = filename_guess
            vendor_source = "filename"
            issues.append(f"vendor name not found in file; inferred {vendor!r} from filename -- verify")
        else:
            issues.append("vendor name not found in file or filename")

    if not date_val:
        issues.append("assessment date not found")

    parsed_date = _coerce_date(date_val)
    if date_val and not parsed_date:
        issues.append(f"could not parse date value: {date_val!r}")

    return HecvatRecord(str(path), ticket_number, vendor, _clean(product), parsed_date, vendor_source, issues)


def parse_directory(dir_path: Path) -> list[HecvatRecord]:
    return [parse_file(p) for p in sorted(dir_path.rglob("*.xlsx")) if not p.name.startswith("~$")]
