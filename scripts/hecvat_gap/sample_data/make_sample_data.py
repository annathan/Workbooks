"""Generates synthetic sample inputs so main.py can be run and sanity-checked
before you have real ServiceNow/Armis exports to point it at. Not part of
the pipeline itself — run once, then delete or ignore its output.
"""
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
ATTACHMENTS = HERE / "hecvat_attachments"
ATTACHMENTS.mkdir(exist_ok=True)


def make_hecvat(filename, vendor, product, date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "General"
    ws["A1"] = "HECVAT Lite - General Information"
    ws["A3"] = "Vendor Legal Name"
    ws["B3"] = vendor
    ws["A4"] = "Product or Service Name"
    ws["B4"] = product
    ws["A5"] = "Date Completed"
    ws["B5"] = date.strftime("%m/%d/%Y")
    wb.save(ATTACHMENTS / filename)


today = datetime(2026, 7, 28)
make_hecvat("RITM0010001_Zoom.xlsx", "Zoom Video Communications", "Zoom Meetings", today - timedelta(days=365 * 3))
make_hecvat("RITM0010002_Slack.xlsx", "Slack Technologies LLC", "Slack", today - timedelta(days=180))
make_hecvat("RITM0010003_Adobe.xlsx", "Adobe Inc.", "Adobe Acrobat", today - timedelta(days=365 * 2.5))

(HERE / "cmdb_export.csv").write_text(
    "Name,Class\n"
    "Zoom,Software\n"
    "Slack,Software\n"
    "Adobe Acrobat,Software\n"
    "Microsoft Teams,Software\n"
    "Google Chrome,Software\n"
)

(HERE / "armis_export.csv").write_text(
    "Application Name,Device Count\n"
    "Zoom,6917\n"
    "Slack,3200\n"
    "Adobe Acrobat,5450\n"
    "Microsoft Teams,5572\n"
    "Google Chrome,6917\n"
)

print(f"Sample data written under {HERE}")
